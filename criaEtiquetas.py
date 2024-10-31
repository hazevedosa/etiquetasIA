from docx import Document
from openpyxl import load_workbook

def fill_existing_word_table(excel_path, word_path, table_index, sheet_name='Planilha1'):
    """
    Fill an existing table in a Word document with Excel data
    table_index: the index of the table to fill (0 for first table, 1 for second, etc.)
    """
    # Load Excel workbook and select sheet
    wb = load_workbook(excel_path)
    ws = wb[sheet_name]
    
    # Load existing Word document
    doc = Document(word_path)
    
    # Get the specified table
    try:
        table = doc.tables[table_index]
    except IndexError:
        print(f"Error: No table found at index {table_index}")
        return
    
    # Get dimensions
    word_rows = len(table.rows)
    word_cols = len(table.columns)
    excel_rows = ws.max_row
    excel_cols = ws.max_column

    offset = 10*table_index
    
    # Fill the table
    for row in range(10):

        if row < 5:
            final_column = 0
        else:
            final_column = 1

        row_ = row + 2

        nome_cell_value = ws.cell(row = offset + row_, column=1).value
        posto_cell_value = ws.cell(row = offset + row_, column=3).value
        instituicao_cell_value = ws.cell(row = offset + row_, column=4).value
        # Convert None to empty string
        nome_cell_value = str(nome_cell_value) if nome_cell_value is not None else ""
        instituicao_cell_value = str(instituicao_cell_value) if instituicao_cell_value is not None else ""
        posto_cell_value = str(posto_cell_value) if posto_cell_value is not None else ""

        table.cell(row - 5*final_column, final_column).text = f"{posto_cell_value}\n{nome_cell_value}\n{instituicao_cell_value}"

    
    # Save the modified document
    doc.save(word_path)
    print(f"Table filled successfully in {word_path}")

# Example usage
if __name__ == "__main__":
    excel_file = "data.xlsx"
    word_file = "etiquetas_teste.docx"
    
    # Fill the first table (index 0) in the Word document
    for i in range(23):
        fill_existing_word_table(excel_file, word_file, table_index=i)
